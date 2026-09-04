import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import pandas as pd

def extract_ncs_master(
    excel_path: str,
    output_csv_path: str
) -> pd.DataFrame:
    """
    NCS 원천 엑셀(24개 대분류 시트)에서 세분류(8자리)까지의 계층 마스터를 추출하여 CSV로 저장.
    """
    print(f"[NCS Master Extraction] Loading Excel from: {excel_path}")
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"File not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    print(f"Total sheets found: {len(sheet_names)}: {sheet_names}")

    seen_codes = set()
    records = []

    for sidx, sname in enumerate(sheet_names, 1):
        ws = wb[sname]
        header = None
        sheet_count = 0
        
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = list(row)
                continue
            
            # 0: 대분류코드, 1: 대분류명, 2: 중분류코드, 3: 중분류명,
            # 4: 소분류코드, 5: 소분류명, 6: 세분류코드, 7: 세분류명
            if len(row) >= 8 and row[0] is not None and row[6] is not None:
                major_cd = str(row[0]).strip().zfill(2)
                mid_cd = str(row[2]).strip().zfill(2)
                minor_cd = str(row[4]).strip().zfill(2)
                detail_sub = str(row[6]).strip().zfill(2)
                
                full_detail_cd = f"{major_cd}{mid_cd}{minor_cd}{detail_sub}"
                
                if full_detail_cd not in seen_codes:
                    seen_codes.add(full_detail_cd)
                    records.append({
                        "major_cd": major_cd,
                        "major_nm": str(row[1]).strip() if row[1] else "",
                        "mid_cd": mid_cd,
                        "mid_nm": str(row[3]).strip() if row[3] else "",
                        "minor_cd": minor_cd,
                        "minor_nm": str(row[5]).strip() if row[5] else "",
                        "detail_cd": full_detail_cd,
                        "detail_nm": str(row[7]).strip() if row[7] else ""
                    })
                    sheet_count += 1
        
        print(f"  [{sidx:2d}/{len(sheet_names)}] Sheet '{sname}': {sheet_count} new unique detail codes")

    wb.close()

    df = pd.DataFrame(records).sort_values("detail_cd").reset_index(drop=True)
    os.makedirs(os.path.dirname(output_csv_path), exist_ok=True)
    df.to_csv(output_csv_path, index=False, encoding="utf-8-sig")
    print(f"\nSuccessfully extracted {len(df)} NCS detail codes.")
    print(f"Saved to: {output_csv_path}")
    return df

if __name__ == "__main__":
    base_dir = r"c:\Users\SD2-21\Desktop\maeng\workspace\SKKULS_AIAGENT\15_MIniProject"
    input_excel = os.path.join(base_dir, "data", "structured", "NCS정보망DB(대분류별,2022년3월).xlsx")
    output_csv = os.path.join(base_dir, "data", "structured", "ncs_codes_master.csv")
    
    df_ncs = extract_ncs_master(input_excel, output_csv)
    print("\nSample records (first 10):")
    print(df_ncs.head(10).to_string())
